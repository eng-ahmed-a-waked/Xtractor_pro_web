import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import os
import logging
from utils.excel_processor import validate_excel_file, convert_xls_to_xlsx

logger = logging.getLogger(__name__)


def style_visits_worksheet(ws, title, creation_datetime):
    """
    Apply professional formatting to visits worksheet with title and date header
    """
    try:
        max_col = ws.max_column
        max_row = ws.max_row
        
        if max_col == 0 or max_row == 0:
            logger.warning("Worksheet is empty, skipping formatting")
            return ws
        
        ws.insert_rows(1, 3)
        
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(1, 1)
        title_cell.value = title
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, size=16, color="FFFFFF", name="Arial")
        title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        date_cell = ws.cell(2, 1)
        date_cell.value = f"📅 تاريخ ووقت الإنشاء: {creation_datetime}"
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
        date_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        table_ref = f"A4:{openpyxl.utils.get_column_letter(max_col)}{max_row + 3}"
        table_name = f"Table_{ws.title[:20].replace(' ', '_').replace('-', '_')}"
        table_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in table_name)
        table_name = table_name[:255]
        
        ws._tables = []
        
        try:
            tab = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
        except Exception as table_error:
            logger.warning(f"Could not apply table to worksheet {ws.title}: {table_error}")
        
        ws.freeze_panes = "A5"
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 10
        ws.row_dimensions[4].height = 25
        
        logger.info(f"Applied formatting to worksheet: {ws.title}")
        return ws
        
    except Exception as e:
        logger.error(f"Error formatting worksheet: {e}")
        return ws


def distribute_visits(file_path, sheet_name, output_folder):
    """
    Distribute visits report by supervisors and representatives
    Creates separate Excel files for each supervisor and representative
    """
    try:
        logger.info(f"Starting visits distribution from: {os.path.basename(file_path)}")
        logger.info(f"Sheet name: {sheet_name}")
        logger.info(f"Output folder: {output_folder}")
        
        # ✅ تحقق من صلاحية الملف أولًا
        is_valid, file_type, error = validate_excel_file(file_path)
        if not is_valid:
            raise Exception(f"Invalid Excel file: {error}")
        
        # ✅ في حال كان XLS أو HTML iTrack يتم تحويله
        if file_type == 'xls':
            logger.info("Converting XLS file to XLSX for compatibility...")
            file_path = convert_xls_to_xlsx(file_path, output_folder)
        
        os.makedirs(output_folder, exist_ok=True)
        
        # ✅ قراءة ملف Excel بعد التحويل إن لزم
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            logger.info(f"Successfully read Excel file with {len(df)} rows")
        except Exception as read_error:
            logger.error(f"Error reading Excel file: {read_error}")
            raise Exception(f"Cannot read Excel file. Please check the file format and sheet name. Error: {str(read_error)}")
        
        # ✅ التحقق من الأعمدة
        if len(df.columns) < 3:
            raise ValueError(f"File must have at least 3 columns (Supervisor, Representative, Line). Found {len(df.columns)} columns.")
        
        df.columns = [str(c).strip() for c in df.columns]
        col_super = df.columns[0]
        col_rep = df.columns[1]
        col_line = df.columns[2]
        
        logger.info(f"Column mapping - Supervisor: {col_super}, Representative: {col_rep}, Line: {col_line}")
        
        df = df.dropna(how='all')
        creation_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        results = {'supervisors': [], 'representatives': [], 'total_files': 0}
        
        # ==================== SUPERVISORS ==================== #
        supervisors = df[col_super].dropna().unique()
        logger.info(f"Found {len(supervisors)} unique supervisors")
        
        for idx, sup in enumerate(supervisors, 1):
            try:
                logger.info(f"Processing supervisor {idx}/{len(supervisors)}: {sup}")
                df_sup = df[df[col_super] == sup].copy()
                if df_sup.empty:
                    continue
                
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
                ws_total = wb.create_sheet(title="إجمالي")
                for r in dataframe_to_rows(df_sup, index=False, header=True):
                    ws_total.append(r)
                
                style_visits_worksheet(ws_total, f"📊 زيارات ومبيعات - {sup}", creation_datetime)
                
                lines = df_sup[col_line].dropna().unique()
                for line in lines:
                    try:
                        sheet_title = str(line)[:31].replace("/", "-").replace("\\", "-").replace(":", "-").replace("*", "-").replace("?", "-").replace("[", "(").replace("]", ")")
                        ws_line = wb.create_sheet(title=sheet_title)
                        df_line = df_sup[df_sup[col_line] == line].copy()
                        for r in dataframe_to_rows(df_line, index=False, header=True):
                            ws_line.append(r)
                        style_visits_worksheet(ws_line, f"📈 زيارات ومبيعات - {line}", creation_datetime)
                    except Exception as e:
                        logger.error(f"Error creating sheet for line {line}: {e}")
                        continue
                
                safe_sup_name = str(sup).replace("/", "-").replace("\\", "-").replace(":", "-").replace("*", "-")[:200]
                filename = f"Supervisor_{safe_sup_name}.xlsx"
                save_path = os.path.join(output_folder, filename)
                wb.save(save_path)
                wb.close()
                
                results['supervisors'].append({
                    'name': str(sup),
                    'filename': filename,
                    'lines_count': len(lines),
                    'records_count': len(df_sup)
                })
            except Exception as e:
                logger.error(f"Error processing supervisor {sup}: {e}")
                continue
        
        # ==================== REPRESENTATIVES ==================== #
        reps = df[col_rep].dropna().unique()
        logger.info(f"Found {len(reps)} unique representatives")
        
        for idx, rep in enumerate(reps, 1):
            try:
                logger.info(f"Processing representative {idx}/{len(reps)}: {rep}")
                df_rep = df[df[col_rep] == rep].copy()
                if df_rep.empty:
                    continue
                
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
                ws_total = wb.create_sheet(title="إجمالي")
                for r in dataframe_to_rows(df_rep, index=False, header=True):
                    ws_total.append(r)
                
                style_visits_worksheet(ws_total, f"📊 زيارات ومبيعات - {rep}", creation_datetime)
                
                lines = df_rep[col_line].dropna().unique()
                for line in lines:
                    try:
                        sheet_title = str(line)[:31].replace("/", "-").replace("\\", "-").replace(":", "-").replace("*", "-").replace("?", "-").replace("[", "(").replace("]", ")")
                        ws_line = wb.create_sheet(title=sheet_title)
                        df_line = df_rep[df_rep[col_line] == line].copy()
                        for r in dataframe_to_rows(df_line, index=False, header=True):
                            ws_line.append(r)
                        style_visits_worksheet(ws_line, f"📈 زيارات ومبيعات - {line}", creation_datetime)
                    except Exception as e:
                        logger.error(f"Error creating sheet for line {line}: {e}")
                        continue
                
                safe_rep_name = str(rep).replace("/", "-").replace("\\", "-").replace(":", "-").replace("*", "-")[:200]
                filename = f"Representative_{safe_rep_name}.xlsx"
                save_path = os.path.join(output_folder, filename)
                wb.save(save_path)
                wb.close()
                
                results['representatives'].append({
                    'name': str(rep),
                    'filename': filename,
                    'lines_count': len(lines),
                    'records_count': len(df_rep)
                })
            except Exception as e:
                logger.error(f"Error processing representative {rep}: {e}")
                continue
        
        results['total_files'] = len(results['supervisors']) + len(results['representatives'])
        logger.info(f"Distribution complete! Supervisors: {len(results['supervisors'])}, Representatives: {len(results['representatives'])}, Total: {results['total_files']}")
        
        return results
        
    except Exception as e:
        logger.error(f"Error distributing visits: {e}", exc_info=True)
        raise

    
def validate_visits_file(file_path, sheet_name="إجمالي"):
    """
    Validate visits file structure before processing
    
    Args:
        file_path (str): Path to the visits Excel file
        sheet_name (str): Name of the sheet to validate
        
    Returns:
        tuple: (is_valid, message, info_dict)
    """
    try:
        if not os.path.exists(file_path):
            return False, "File does not exist", {}
        
        # Try to read the file
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        except Exception as read_error:
            return False, f"Cannot read Excel file: {str(read_error)}", {}
        
        # Check minimum columns
        if len(df.columns) < 3:
            return False, f"File must have at least 3 columns. Found {len(df.columns)}", {}
        
        # Get column info
        col_super = df.columns[0]
        col_rep = df.columns[1]
        col_line = df.columns[2]
        
        # Count unique values
        supervisors_count = df[col_super].dropna().nunique()
        reps_count = df[col_rep].dropna().nunique()
        lines_count = df[col_line].dropna().nunique()
        total_rows = len(df)
        
        info = {
            'total_rows': total_rows,
            'supervisors_count': supervisors_count,
            'representatives_count': reps_count,
            'lines_count': lines_count,
            'columns': list(df.columns)
        }
        
        if supervisors_count == 0:
            return False, "No supervisors found in first column", info
        
        if reps_count == 0:
            return False, "No representatives found in second column", info
        
        return True, "File is valid and ready for processing", info
        
    except Exception as e:
        return False, f"Validation error: {str(e)}", {}
    

def get_distribution_summary(results):
    """
    Generate a human-readable summary of distribution results
    
    Args:
        results (dict): Results dictionary from distribute_visits
        
    Returns:
        str: Formatted summary text
    """
    try:
        summary_lines = [
            "=" * 70,
            "📊 VISITS DISTRIBUTION SUMMARY",
            "=" * 70,
            "",
            f"✅ Total Files Created: {results.get('total_files', 0)}",
            "",
            f"👔 Supervisors: {len(results.get('supervisors', []))}",
        ]
        
        for sup in results.get('supervisors', []):
            summary_lines.append(f"  • {sup['name']}: {sup['lines_count']} lines, {sup['records_count']} records")
        
        summary_lines.append("")
        summary_lines.append(f"👥 Representatives: {len(results.get('representatives', []))}")
        
        for rep in results.get('representatives', []):
            summary_lines.append(f"  • {rep['name']}: {rep['lines_count']} lines, {rep['records_count']} records")
        
        summary_lines.append("")
        summary_lines.append("=" * 70)
        
        return "\n".join(summary_lines)
        
    except Exception as e:
        return f"Error generating summary: {e}"

    
