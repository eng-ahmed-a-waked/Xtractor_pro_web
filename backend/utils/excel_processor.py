import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import os
import logging

logger = logging.getLogger(__name__)


# ======================== تحويل الملفات المحسّن ======================== #

def convert_xls_to_xlsx(file_path, output_folder):
    """
    تحويل ملف XLS إلى XLSX بشكل احترافي
    يدعم ملفات XLS الحقيقية وملفات HTML من iTrack
    
    Args:
        file_path: مسار الملف المصدر
        output_folder: مجلد الحفظ
    
    Returns:
        str: مسار الملف المحول
    """
    try:
        filename = os.path.basename(file_path)
        name, ext = os.path.splitext(filename)
        
        if ext.lower() != '.xls':
            logger.info(f"File {filename} is already in XLSX format")
            return file_path
        
        logger.info(f"Converting XLS file: {filename}")
        
        df_sheets = {}
        
        try:
            # المحاولة الأولى: قراءة XLS حقيقي
            xls_file = pd.ExcelFile(file_path, engine='xlrd')
            for sheet_name in xls_file.sheet_names:
                df_sheets[sheet_name] = pd.read_excel(xls_file, sheet_name=sheet_name, header=None)
            xls_file.close()
            logger.info(f"Successfully read as XLS: {filename}")
            
        except Exception as e:
            # المحاولة الثانية: قراءة HTML (iTrack)
            logger.warning(f"Standard XLS read failed ({e}), trying HTML fallback...")
            
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
            
            if "<table" in content.lower():
                html_sheets = pd.read_html(content)
                if not html_sheets:
                    raise Exception("No <table> elements found in HTML content")
                df_sheets["iTrack Report"] = html_sheets[0]
                logger.info(f"Successfully parsed HTML iTrack report: {filename}")
            else:
                raise Exception(f"Unsupported file format: {e}")
        
        # تصدير إلى XLSX باستخدام openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # حذف الشيت الافتراضي
        
        for sheet_name, df in df_sheets.items():
            # اسم آمن للشيت (32 حرف كحد أقصى)
            safe_name = sheet_name[:31]
            ws = wb.create_sheet(title=safe_name)
            
            # كتابة البيانات
            for r_idx, row in enumerate(df.values, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = None if pd.isna(value) else str(value)
            
            # إضافة تنسيق جدول
            try:
                max_row, max_col = ws.max_row, ws.max_column
                if max_row > 0 and max_col > 0:
                    # إنشاء اسم جدول فريد
                    table_name = f"Table_{safe_name.replace(' ', '_')}"
                    table_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in table_name)
                    table_name = table_name[:255]  # حد Excel
                    
                    # نطاق الجدول
                    table_range = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}"
                    
                    # إنشاء الجدول
                    tab = Table(displayName=table_name, ref=table_range)
                    style = TableStyleInfo(
                        name="TableStyleMedium9",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False
                    )
                    tab.tableStyleInfo = style
                    ws.add_table(tab)
                    
                    logger.info(f"Applied table formatting to sheet: {safe_name}")
                    
            except Exception as table_error:
                logger.warning(f"Could not apply table formatting to {sheet_name}: {table_error}")
        
        # حفظ الملف
        output_path = os.path.join(output_folder, f"{name}.xlsx")
        wb.save(output_path)
        wb.close()
        
        if os.path.exists(output_path):
            logger.info(f"Successfully converted to XLSX: {output_path}")
            return output_path
        else:
            raise Exception("Output file not created")
    
    except Exception as e:
        logger.error(f"Error converting file {file_path}: {e}", exc_info=True)
        raise Exception(f"Conversion failed: {str(e)}")


def validate_excel_file(file_path):
    """
    التحقق من صلاحية ملف Excel
    
    Args:
        file_path: مسار الملف
    
    Returns:
        tuple: (is_valid, file_type, error_message)
    """
    try:
        if not os.path.exists(file_path):
            return False, None, "File does not exist"
        
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext not in ['.xls', '.xlsx']:
            return False, None, "File is not an Excel file (.xls or .xlsx)"
        
        # محاولة الفتح
        if file_ext == '.xls':
            try:
                pd.ExcelFile(file_path, engine='xlrd')
                return True, 'xls', None
            except Exception as e:
                return False, 'xls', f"Invalid XLS file: {str(e)}"
        
        else:  # .xlsx
            try:
                openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                return True, 'xlsx', None
            except Exception as e:
                return False, 'xlsx', f"Invalid XLSX file: {str(e)}"
        
    except Exception as e:
        return False, None, f"Validation error: {str(e)}"


def batch_convert_xls_files(folder_path, output_folder=None):
    """
    تحويل جميع ملفات XLS في مجلد
    
    Args:
        folder_path: مسار المجلد
        output_folder: مجلد الحفظ (اختياري)
    
    Returns:
        dict: نتائج التحويل
    """
    if output_folder is None:
        output_folder = folder_path
    
    results = {
        'converted': [],
        'errors': [],
        'skipped': [],
        'total': 0
    }
    
    try:
        # التأكد من وجود مجلد الحفظ
        os.makedirs(output_folder, exist_ok=True)
        
        # البحث عن ملفات Excel
        for filename in os.listdir(folder_path):
            if filename.startswith('~$') or filename.startswith('.'):
                continue
            
            file_path = os.path.join(folder_path, filename)
            
            if not os.path.isfile(file_path):
                continue
            
            file_ext = os.path.splitext(filename)[1].lower()
            
            if file_ext not in ['.xls', '.xlsx']:
                continue
            
            results['total'] += 1
            
            try:
                if file_ext == '.xls':
                    # تحويل XLS إلى XLSX
                    converted_path = convert_xls_to_xlsx(file_path, output_folder)
                    results['converted'].append({
                        'original': filename,
                        'converted': os.path.basename(converted_path)
                    })
                    logger.info(f"Converted: {filename}")
                else:
                    # بالفعل XLSX
                    results['skipped'].append(filename)
                    logger.info(f"Skipped (already XLSX): {filename}")
                    
            except Exception as e:
                results['errors'].append({
                    'file': filename,
                    'error': str(e)
                })
                logger.error(f"Failed to convert {filename}: {e}")
        
        logger.info(f"Batch conversion complete. Converted: {len(results['converted'])}, "
                   f"Errors: {len(results['errors'])}, Skipped: {len(results['skipped'])}")
        return results
        
    except Exception as e:
        logger.error(f"Error in batch conversion: {e}")
        raise


def format_worksheet_with_header(ws, title, creation_datetime):
    """
    تطبيق تنسيق احترافي على الورقة مع العنوان والتاريخ
    
    Args:
        ws: ورقة العمل
        title: العنوان
        creation_datetime: التاريخ والوقت
    
    Returns:
        الورقة المنسقة
    """
    try:
        max_col = ws.max_column
        max_row = ws.max_row
        
        if max_col == 0 or max_row == 0:
            logger.warning("Worksheet is empty, skipping formatting")
            return ws
        
        # إدراج 3 صفوف في الأعلى
        ws.insert_rows(1, 3)
        
        # صف العنوان (الصف 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(1, 1)
        title_cell.value = title
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, size=16, color="FFFFFF", name="Arial")
        title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        # صف التاريخ (الصف 2)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        date_cell = ws.cell(2, 1)
        date_cell.value = f"📅 تاريخ ووقت الإنشاء: {creation_datetime}"
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
        date_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # الصف 3 فارغ (فاصل)
        # الصف 4 سيحتوي على الهيدر الأصلي
        
        # ضبط ارتفاعات الصفوف
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 10
        ws.row_dimensions[4].height = 25
        
        # إضافة حدود لصفوف العنوان والتاريخ
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for row in [1, 2]:
            for col in range(1, max_col + 1):
                ws.cell(row, col).border = thin_border
        
        logger.info(f"Applied header formatting to worksheet: {ws.title}")
        return ws
        
    except Exception as e:
        logger.error(f"Error formatting worksheet header: {e}")
        return ws


def apply_professional_formatting(ws):
    """
    تطبيق التنسيق الاحترافي الكامل على ورقة Excel
    
    Args:
        ws: ورقة العمل
    """
    try:
        # التنسيقات الأساسية
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        
        data_fill_1 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        data_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # تنسيق الهيدر (الصف الأول بعد العنوان)
        header_row = 4  # بعد صفوف العنوان والتاريخ والفاصل
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(header_row, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # تنسيق صفوف البيانات (تبديل الألوان)
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), 1):
            fill = data_fill_1 if row_idx % 2 == 0 else data_fill_2
            
            for cell in row:
                cell.border = border
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # ضبط عرض الأعمدة تلقائيًا
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # حد أقصى 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Applied professional formatting to worksheet: {ws.title}")
        
    except Exception as e:
        logger.error(f"Error applying professional formatting: {e}")


# ======================== دوال إضافية للتحسين ======================== #

def add_excel_table(ws, start_row=4):
    """
    إضافة جدول Excel احترافي
    
    Args:
        ws: ورقة العمل
        start_row: رقم صف البداية (بعد العنوان)
    """
    try:
        max_col = ws.max_column
        max_row = ws.max_row
        
        if max_row <= start_row:
            return
        
        # نطاق الجدول
        table_ref = f"A{start_row}:{openpyxl.utils.get_column_letter(max_col)}{max_row}"
        
        # اسم الجدول
        table_name = f"Table_{ws.title[:20].replace(' ', '_')}"
        table_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in table_name)
        table_name = table_name[:255]
        
        # حذف الجداول القديمة
        ws._tables = []
        
        # إنشاء الجدول
        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        # تجميد الصفوف
        ws.freeze_panes = f"A{start_row + 1}"
        
        logger.info(f"Added Excel table to worksheet: {ws.title}")
        
    except Exception as e:
        logger.warning(f"Could not add table to worksheet: {e}")


def optimize_workbook(wb):
    """
    تحسين المصنف بالكامل
    
    Args:
        wb: المصنف
    """
    try:
        # إزالة الشيتات الفارغة
        for ws in wb.worksheets:
            if ws.max_row == 0 or ws.max_column == 0:
                wb.remove(ws)
        
        # ضبط الخصائص
        wb.properties.title = "Xtractor Pro Report"
        wb.properties.creator = "Xtractor Pro v3.0"
        wb.properties.description = "Professional Excel Report with Enhanced Formatting"
        
        logger.info("Workbook optimized successfully")
        
    except Exception as e:
        logger.warning(f"Error optimizing workbook: {e}")
