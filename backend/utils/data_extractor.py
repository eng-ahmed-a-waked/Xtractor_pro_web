import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import re
import os
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


# ======================== استخراج البيانات المحسّن ======================== #

def extract_car_number(car_code):
    """استخراج رقم السيارة من الكود (أول رقم في السلسلة)"""
    if not car_code:
        return ""
    try:
        car_str = str(car_code).strip()
        numbers = re.findall(r'\d+', car_str)
        return numbers[0] if numbers else car_str
    except Exception as e:
        logger.warning(f"Error extracting car number from {car_code}: {e}")
        return str(car_code) if car_code else ""


def extract_car_plate(car_code):
    """استخراج أرقام اللوحة فقط (جميع الأرقام مدمجة)"""
    if not car_code:
        return ""
    try:
        text = str(car_code).strip()
        numbers = ''.join(re.findall(r'\d+', text))
        return numbers if numbers else text
    except Exception as e:
        logger.warning(f"Error extracting car plate from {car_code}: {e}")
        return str(car_code) if car_code else ""


def extract_coordinates(text):
    """استخراج الإحداثيات الرقمية فقط (lat,lon)"""
    if not text:
        return ""
    try:
        text_str = str(text).strip()
        
        # محاولة أنماط مختلفة للإحداثيات
        patterns = [
            r'(\d+\.\d+)[NS]?\s*,\s*(\d+\.\d+)[EW]?',  # مع الاتجاه
            r'(\d+\.\d+)\s*,\s*(\d+\.\d+)',  # بدون اتجاه
            r'(\d+\.\d+)\s+(\d+\.\d+)'  # بفراغ
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text_str)
            if match:
                lat, lon = match.groups()
                return f"{lat},{lon}"
                
    except Exception as e:
        logger.warning(f"Error extracting coordinates from {text}: {e}")
    
    return ""


def extract_address_text(text):
    """استخراج النص الوصفي للموقع (بدون الإحداثيات)"""
    if not text:
        return ""
    
    try:
        text_str = str(text).strip()
        
        if not text_str:
            return ""
        
        # إزالة جميع أنماط الإحداثيات
        coordinate_patterns = [
            r'\b\d+\.\d+[NS]?\s*,\s*\d+\.\d+[EW]?\b',
            r'\b\d+\.\d+\s*,\s*\d+\.\d+\b',
            r'\b\d+\.\d+,\d+\.\d+\b',
            r'\b\d+\.\d+\s+\d+\.\d+\b'
        ]
        
        for pattern in coordinate_patterns:
            text_str = re.sub(pattern, '', text_str)
        
        # تنظيف الفواصل والمسافات الزائدة
        text_str = re.sub(r',\s*,+', ',', text_str)
        text_str = re.sub(r'\s+', ' ', text_str)
        text_str = text_str.strip(', ')
        
        # إزالة حروف الاتجاه المنفصلة
        text_str = re.sub(r'\s+[NSEW]\s+', ' ', text_str)
        text_str = re.sub(r'^[NSEW]\s+', '', text_str)
        text_str = re.sub(r'\s+[NSEW]$', '', text_str)
        
        text_str = text_str.strip()
        
        # إرجاع النص إذا كان ذو معنى
        if len(text_str) > 3:
            return text_str
        
        return ""
        
    except Exception as e:
        logger.warning(f"Error extracting address text from {text}: {e}")
        return ""


def safe_read_cell(cell):
    """قراءة آمنة لخلية Excel"""
    try:
        if cell.value is None:
            return None
        if isinstance(cell.value, datetime):
            return cell.value
        if isinstance(cell.value, (int, float)):
            return cell.value
        return str(cell.value).strip()
    except Exception as e:
        logger.warning(f"Error reading cell: {e}")
        return None


def check_zone(coordinates, zone_points=None):
    """التحقق من أن الإحداثيات داخل النطاق المحدد"""
    if not coordinates:
        return "غير محدد"
    
    try:
        coord_str = str(coordinates).strip()
        
        if ',' not in coord_str:
            return "غير محدد"
        
        lat, lon = map(float, coord_str.split(','))
        
        # النقاط الافتراضية للمنطقة
        if zone_points is None:
            zone_points = [
                (30.22923, 31.73212),
                (30.22984, 31.73347),
                (30.22908, 31.73418),
                (30.22851, 31.73325)
            ]
        
        # فحص بسيط باستخدام bounding box
        lat_min = min(p[0] for p in zone_points)
        lat_max = max(p[0] for p in zone_points)
        lon_min = min(p[1] for p in zone_points)
        lon_max = max(p[1] for p in zone_points)
        
        if lat_min <= lat <= lat_max and lon_min <= lon <= lon_max:
            return "داخل النطاق"
        else:
            return "خارج النطاق"
            
    except Exception as e:
        logger.warning(f"Error checking zone for coordinates {coordinates}: {e}")
        return "غير محدد"


def extract_data_from_excel(file_path, mode="engine_idle", zone_points=None):
    """
    استخراج البيانات من ملف Excel واحد
    
    Args:
        file_path: مسار الملف
        mode: وضع الاستخراج ('engine_idle' أو 'parking_details')
        zone_points: نقاط حدود المنطقة
    
    Returns:
        list: قائمة السجلات المستخرجة
    """
    extracted_data = []
    
    try:
        logger.info(f"Starting extraction from: {os.path.basename(file_path)}")
        
        # فتح الملف
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        
        for ws in wb.worksheets:
            try:
                sheet_title = ws.title
                logger.info(f"Processing sheet: {sheet_title}")
                
                # استخراج كود السيارة من A1
                car_code_cell = ws['A1']
                car_code = safe_read_cell(car_code_cell)
                
                if car_code:
                    if mode == "engine_idle":
                        car_code = extract_car_number(car_code)
                    else:
                        car_code = extract_car_plate(car_code)
                
                # البحث عن صف الهيدر
                header_row = None
                columns = {
                    'start': None,
                    'end': None,
                    'duration': None,
                    'address': None,
                    'coordinate': None
                }
                
                # الكلمات المفتاحية للبحث
                keywords = {
                    'start': ['start time', 'بداية', 'start', 'وقت البدء'],
                    'end': ['end time', 'نهاية', 'end', 'وقت النهاية'],
                    'duration': ['duration', 'مدة', 'stop duration'],
                    'address': ['address', 'موقع', 'عنوان'],
                    'coordinate': ['coordinate', 'إحداثيات', 'احداثيات']
                }
                
                # البحث في أول 20 صف
                for row_idx in range(1, min(21, ws.max_row + 1)):
                    row_values = [str(safe_read_cell(cell)).lower() if safe_read_cell(cell) else '' 
                                  for cell in ws[row_idx]]
                    row_str = ' '.join(row_values)
                    
                    # التحقق من وجود كلمات مفتاحية
                    has_start = any(kw in row_str for kw in keywords['start'])
                    has_end = any(kw in row_str for kw in keywords['end'])
                    has_duration = any(kw in row_str for kw in keywords['duration'])
                    
                    if has_start or (has_end and has_duration):
                        header_row = row_idx
                        
                        # تحديد مواقع الأعمدة
                        for col_idx, cell in enumerate(ws[row_idx], 1):
                            cell_val = safe_read_cell(cell)
                            if not cell_val:
                                continue
                            
                            cell_str = str(cell_val).lower().strip()
                            
                            if not columns['start'] and any(kw in cell_str for kw in keywords['start']):
                                columns['start'] = col_idx
                            elif not columns['end'] and any(kw in cell_str for kw in keywords['end']):
                                columns['end'] = col_idx
                            elif not columns['duration'] and any(kw in cell_str for kw in keywords['duration']):
                                columns['duration'] = col_idx
                            elif not columns['address'] and any(kw in cell_str for kw in keywords['address']):
                                columns['address'] = col_idx
                            elif not columns['coordinate'] and any(kw in cell_str for kw in keywords['coordinate']):
                                columns['coordinate'] = col_idx
                        
                        break
                
                if not header_row:
                    logger.warning(f"No header row found in sheet: {sheet_title}")
                    continue
                
                if not any([columns['start'], columns['end'], columns['duration']]):
                    logger.warning(f"No time columns found in sheet: {sheet_title}")
                    continue
                
                # استخراج البيانات
                records_count = 0
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    try:
                        row = list(ws[row_idx])
                        
                        # تخطي الصفوف الفارغة
                        if not any(safe_read_cell(cell) for cell in row):
                            continue
                        
                        # قراءة البيانات
                        start_time = safe_read_cell(row[columns['start'] - 1]) if columns['start'] else None
                        end_time = safe_read_cell(row[columns['end'] - 1]) if columns['end'] else None
                        duration = safe_read_cell(row[columns['duration'] - 1]) if columns['duration'] else None
                        address = safe_read_cell(row[columns['address'] - 1]) if columns['address'] else None
                        coordinate = safe_read_cell(row[columns['coordinate'] - 1]) if columns['coordinate'] else None
                        
                        # استخراج الإحداثيات والعنوان
                        numeric_coordinates = extract_coordinates(coordinate) or extract_coordinates(address)
                        address_text = extract_address_text(address) or extract_address_text(coordinate)
                        
                        # التحقق من النطاق
                        zone_status = check_zone(numeric_coordinates, zone_points)
                        
                        # إضافة السجل إذا كان يحتوي على بيانات
                        if start_time or end_time or duration:
                            record = {
                                'car_code': car_code or "",
                                'start_time': start_time,
                                'end_time': end_time,
                                'duration': duration,
                                'coordinates': numeric_coordinates,
                                'zone': zone_status,
                                'address': address_text,
                                'source_sheet': sheet_title
                            }
                            
                            extracted_data.append(record)
                            records_count += 1
                    
                    except Exception as row_error:
                        logger.warning(f"Error processing row {row_idx}: {row_error}")
                        continue
                
                if records_count > 0:
                    logger.info(f"Extracted {records_count} records from sheet: {sheet_title}")
            
            except Exception as sheet_error:
                logger.error(f"Error processing sheet '{ws.title}': {sheet_error}")
                continue
        
        wb.close()
        logger.info(f"Total records extracted: {len(extracted_data)}")
        
    except Exception as e:
        logger.error(f"Error extracting data from {file_path}: {e}", exc_info=True)
        raise
    
    return extracted_data


def get_extraction_statistics(data):
    """
    حساب إحصائيات البيانات المستخرجة
    
    Args:
        data: قائمة السجلات
    
    Returns:
        dict: الإحصائيات
    """
    try:
        stats = {
            'total_records': len(data),
            'inside_zone': 0,
            'outside_zone': 0,
            'undefined_zone': 0,
            'unique_cars': set(),
            'sheets_processed': set()
        }
        
        for record in data:
            zone = record.get('zone', 'غير محدد')
            
            if zone == 'داخل النطاق':
                stats['inside_zone'] += 1
            elif zone == 'خارج النطاق':
                stats['outside_zone'] += 1
            else:
                stats['undefined_zone'] += 1
            
            if record.get('car_code'):
                stats['unique_cars'].add(record['car_code'])
            
            if record.get('source_sheet'):
                stats['sheets_processed'].add(record['source_sheet'])
        
        stats['unique_cars'] = len(stats['unique_cars'])
        stats['sheets_processed'] = len(stats['sheets_processed'])
        
        return stats
        
    except Exception as e:
        logger.error(f"Error calculating statistics: {e}")
        return {
            'total_records': len(data),
            'inside_zone': 0,
            'outside_zone': 0,
            'undefined_zone': 0,
            'unique_cars': 0,
            'sheets_processed': 0
        }


def create_summary_report(data, output_path, mode="engine_idle"):
    """
    إنشاء تقرير موحّد بتنسيق احترافي
    
    Args:
        data: قائمة السجلات
        output_path: مسار الحفظ
        mode: وضع التقرير
    
    Returns:
        str: مسار الملف المحفوظ
    """
    if not data:
        raise ValueError("No data to create report")
    
    try:
        logger.info(f"Creating summary report with {len(data)} records")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "التقرير الموحد"
        
        # ======================== التنسيقات ======================== #
        
        # ألوان الهيدر
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        
        # ألوان النطاق
        inside_zone_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        outside_zone_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        undefined_zone_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # الحدود
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # ======================== صف العنوان ======================== #
        
        car_header = "كود السيارة" if mode == "engine_idle" else "لوحة السيارة"
        headers = ['#', car_header, 'بداية التوقف', 'نهاية التوقف', 'مدة التوقف', 
                   'الإحداثيات', 'النطاق', 'الموقع', 'المصدر']
        
        # دمج الخلايا للعنوان
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_text = "تقرير توقفات السيارات" if mode == "engine_idle" else "تقرير مواقع السيارات"
        title_cell.value = f"🚗 {title_text} - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(name='Arial', size=14, bold=True, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        ws.row_dimensions[1].height = 30
        
        # ======================== صف الهيدر ======================== #
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        ws.row_dimensions[2].height = 25
        
        # ======================== صفوف البيانات ======================== #
        
        for idx, record in enumerate(data, 1):
            row_num = idx + 2
            
            # الرقم التسلسلي
            cell = ws.cell(row=row_num, column=1)
            cell.value = idx
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # كود السيارة
            cell = ws.cell(row=row_num, column=2)
            cell.value = record.get('car_code', '')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            cell.font = Font(name='Arial', size=11, bold=True)
            
            # وقت البداية
            cell = ws.cell(row=row_num, column=3)
            start_val = record.get('start_time', '')
            if isinstance(start_val, datetime):
                cell.value = start_val
                cell.number_format = 'DD/MM/YYYY HH:MM:SS'
            else:
                cell.value = start_val
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # وقت النهاية
            cell = ws.cell(row=row_num, column=4)
            end_val = record.get('end_time', '')
            if isinstance(end_val, datetime):
                cell.value = end_val
                cell.number_format = 'DD/MM/YYYY HH:MM:SS'
            else:
                cell.value = end_val
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # المدة
            cell = ws.cell(row=row_num, column=5)
            cell.value = record.get('duration', '')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            cell.font = Font(name='Arial', size=11, bold=True, color="C00000")
            
            # الإحداثيات (مع رابط Google Maps)
            cell = ws.cell(row=row_num, column=6)
            coordinates = record.get('coordinates', '')
            cell.value = coordinates
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            if coordinates:
                maps_url = f'https://www.google.com/maps?q={coordinates}'
                cell.hyperlink = maps_url
                cell.font = Font(name='Arial', size=10, color="0563C1", underline="single", bold=True)
            
            # النطاق (مع تلوين)
            cell = ws.cell(row=row_num, column=7)
            zone_status = record.get('zone', 'غير محدد')
            cell.value = zone_status
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            if zone_status == "داخل النطاق":
                cell.fill = inside_zone_fill
                cell.font = Font(name='Arial', size=11, bold=True, color="006100")
            elif zone_status == "خارج النطاق":
                cell.fill = outside_zone_fill
                cell.font = Font(name='Arial', size=11, bold=True, color="9C0006")
            else:
                cell.fill = undefined_zone_fill
                cell.font = Font(name='Arial', size=11, bold=True, color="9C6500")
            
            # العنوان
            cell = ws.cell(row=row_num, column=8)
            address_text = record.get('address', '')
            
            if not address_text:
                cell.value = "غير متوفر"
                cell.font = Font(name='Arial', size=10, color="999999", italic=True)
            else:
                cell.value = address_text
                cell.font = Font(name='Arial', size=10, bold=True, color="0066CC")
            
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border = border
            
            # المصدر
            cell = ws.cell(row=row_num, column=9)
            cell.value = record.get('source_sheet', '')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            cell.font = Font(name='Arial', size=9, color="666666")
        
        # ======================== ضبط العرض ======================== #
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 50
        ws.column_dimensions['I'].width = 30
        
        # تجميد الصفوف
        ws.freeze_panes = 'A3'
        
        # حفظ الملف
        wb.save(output_path)
        wb.close()
        
        logger.info(f"Summary report created successfully: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"Error creating summary report: {e}", exc_info=True)
        raise
